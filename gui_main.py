import logging
import os
import re
import threading
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import ttk, filedialog, messagebox

import ddddocr
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright


class ConfigGUI:

    def __init__(self, master):
        self.headless_var = None
        self.export_path_btn = None
        self.export_path_entry = None
        self.export_var = None
        self.run_btn = None
        self.overwrite_var = None
        self.max_row_entry = None
        self.start_row_entry = None
        self.column_entries = {}
        self.url_entry = None
        self.excel_path_entry = None
        self.master = master
        self.scraper_thread = None
        self.create_widgets()
        self.setup_defaults()

    def create_widgets(self):

        self.master.title("成绩查询系统配置")
        self.master.geometry("520x420")

        config_frame = ttk.LabelFrame(self.master, text="系统配置")
        config_frame.pack(padx=15, pady=10, fill="both", expand=True)

        ttk.Label(config_frame, text="Excel文件路径:").grid(row=0, column=0, sticky="w", pady=2)
        self.excel_path_entry = ttk.Entry(config_frame, width=35)
        self.excel_path_entry.grid(row=0, column=1, padx=5)
        ttk.Button(config_frame, text="浏览", command=self.select_excel).grid(row=0, column=2)

        ttk.Label(config_frame, text="查询网址:").grid(row=1, column=0, sticky="w", pady=2)
        self.url_entry = ttk.Entry(config_frame, width=35)
        self.url_entry.grid(row=1, column=1, columnspan=2, sticky="ew", padx=5)

        columns_config = [
            ("姓名列", 2), ("准考证号列", 3),
            ("密码列", 5), ("成绩起始列", 6)
        ]
        for idx, (label, default) in enumerate(columns_config, start=2):
            ttk.Label(config_frame, text=f"{label}:").grid(row=idx, column=0, sticky="w", pady=2)
            entry = ttk.Entry(config_frame, width=5)
            entry.insert(0, str(default))
            entry.grid(row=idx, column=1, sticky="w", padx=5)
            self.column_entries[label] = entry

        ttk.Label(config_frame, text="起始行:").grid(row=6, column=0, sticky="w", pady=2)
        self.start_row_entry = ttk.Entry(config_frame, width=5)
        self.start_row_entry.insert(0, "2")
        self.start_row_entry.grid(row=6, column=1, sticky="w", padx=5)

        ttk.Label(config_frame, text="结束行:").grid(row=7, column=0, sticky="w", pady=2)
        self.max_row_entry = ttk.Entry(config_frame, width=5)
        self.max_row_entry.insert(0, "271")
        self.max_row_entry.grid(row=7, column=1, sticky="w", padx=5)

        self.overwrite_var = tk.BooleanVar()
        ttk.Checkbutton(config_frame, text="覆盖已有成绩", variable=self.overwrite_var).grid(
            row=8, columnspan=3, pady=2)

        self.export_var = tk.BooleanVar()
        ttk.Checkbutton(config_frame, text="导出到新Excel文件", variable=self.export_var,
                        command=self.toggle_export).grid(row=9, columnspan=3, pady=2)

        ttk.Label(config_frame, text="导出路径:").grid(row=10, column=0, sticky="w", pady=2)
        self.export_path_entry = ttk.Entry(config_frame, width=35)
        self.export_path_entry.grid(row=10, column=1, padx=5)
        self.export_path_btn = ttk.Button(config_frame, text="浏览", command=self.select_export_path, state="disabled")
        self.export_path_btn.grid(row=10, column=2)

        self.headless_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(config_frame, text="显示浏览器界面", variable=self.headless_var).grid(row=11, columnspan=3, pady=10)

        btn_frame = ttk.Frame(self.master)
        btn_frame.pack(pady=15)
        self.run_btn = ttk.Button(btn_frame, text="开始运行", command=self.start_scraper)
        self.run_btn.pack(side="left", padx=10)
        ttk.Button(btn_frame, text="退出", command=self.master.quit).pack(side="left")

    def toggle_export(self):

        if self.export_var.get():
            self.export_path_entry.config(state="normal")
            self.export_path_btn.config(state="normal")
        else:
            self.export_path_entry.config(state="disabled")
            self.export_path_btn.config(state="disabled")

    def setup_defaults(self):

        self.url_entry.insert(0, "https://cx.shmeea.edu.cn/shmeea/q/hgk2025yswquery3zd8#")
        self.excel_path_entry.insert(0, "./excel/报名号.xlsx")

    def select_excel(self):

        path = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx")])
        if path:
            self.excel_path_entry.delete(0, tk.END)
            self.excel_path_entry.insert(0, path)

    def select_export_path(self):

        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel文件", "*.xlsx")])
        if path:
            self.export_path_entry.delete(0, tk.END)
            self.export_path_entry.insert(0, path)

    def validate_inputs(self):

        try:

            int(self.start_row_entry.get())
            int(self.max_row_entry.get())
            for entry in self.column_entries.values():
                int(entry.get())

            if self.export_var.get() and not self.export_path_entry.get():
                messagebox.showerror("输入错误", "请选择导出路径")
                return False
            return True
        except ValueError:
            messagebox.showerror("输入错误", "请输入有效的数字")
            return False

    def get_config(self):

        return {
            "excel_path": self.excel_path_entry.get(),
            "target_url": self.url_entry.get(),
            "column_mapping": {
                "name": int(self.column_entries["姓名列"].get()),
                "exam_id": int(self.column_entries["准考证号列"].get()),
                "password": int(self.column_entries["密码列"].get()),
                "first_score": int(self.column_entries["成绩起始列"].get())
            },
            "start_row": int(self.start_row_entry.get()),
            "max_row": int(self.max_row_entry.get()),
            "overwrite": self.overwrite_var.get(),
            "export_new": self.export_var.get(),
            "export_path": self.export_path_entry.get(),
            "headless": not self.headless_var.get()
        }

    def start_scraper(self):

        if not self.validate_inputs():
            return

        config = self.get_config()
        self.run_btn.config(state="disabled")

        def run_thread():
            try:
                scraper = ScoreScraper(config)
                scraper.run()
                messagebox.showinfo("完成", "所有考生成绩处理完成！")
            except Exception as e:
                messagebox.showerror("错误", f"运行失败: {str(e)}")
            finally:
                self.run_btn.config(state="normal")

        self.scraper_thread = threading.Thread(target=run_thread, daemon=True)
        self.scraper_thread.start()


class ScoreScraper:

    def __init__(self, user_config):
        self.user_config = user_config
        self.config = self.get_merged_config()
        self.ocr = ddddocr.DdddOcr()
        self._setup_directories()
        self._setup_logging()
        self.logger = logging.getLogger(__name__)
        self.workbook = None

    def get_merged_config(self):

        default_dirs = {
            "screenshot_dir": Path("./verify"),
            "html_temp_dir": Path("./html"),
            "selected_html_dir": Path("./selected_html"),
            "scores_dir": Path("./scores"),
            "log_dir": Path("./logs"),
        }
        return {
            **default_dirs,
            "excel_path": Path(self.user_config["excel_path"]),
            "target_url": self.user_config["target_url"],
            "column_mapping": self.user_config["column_mapping"],
            "start_row": self.user_config["start_row"],
            "max_row": self.user_config["max_row"],
            "overwrite": self.user_config["overwrite"],
            "export_new": self.user_config["export_new"],
            "export_path": self.user_config["export_path"],
            "headless": self.user_config["headless"]
        }

    def _setup_directories(self):

        for dir_key in ["screenshot_dir", "html_temp_dir",
                        "selected_html_dir", "scores_dir", "log_dir"]:
            self.config[dir_key].mkdir(parents=True, exist_ok=True)

    def _setup_logging(self):

        log_file = self.config["log_dir"] / f"system_{datetime.now().strftime('%Y%m%d_%H%M')}.log"
        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
            handlers=[
                logging.FileHandler(log_file, encoding="utf-8"),
                logging.StreamHandler()
            ]
        )

    def _extract_scores(self, html_path: Path) -> list:

        score_pattern = re.compile(r'>(\d+\.?\d*)<')
        scores = []
        try:
            with open(html_path, 'r', encoding='utf-8') as f:
                for line in f:
                    scores.extend(score_pattern.findall(line))
            self.logger.debug("成功提取 %d 个成绩", len(scores))
        except Exception as e:
            self.logger.error("提取失败: %s", str(e))
        return scores

    def _process_student(self, row: int):

        ws = self.workbook.active
        exam_id = str(ws.cell(row=row, column=self.config["column_mapping"]["exam_id"]).value)
        password = str(ws.cell(row=row, column=self.config["column_mapping"]["password"]).value)
        name = str(ws.cell(row=row, column=self.config["column_mapping"]["name"]).value)

        if not password:
            self.logger.warning("跳过 %s：无密码", name)
            return

        with sync_playwright() as playwright:
            try:
                browser = playwright.chromium.launch(headless=self.config["headless"])
                context = browser.new_context()
                page = context.new_page()

                # 网页操作流程
                page.goto(self.config["target_url"])
                page.wait_for_timeout(500)

                page.locator("input[name=\"BMH\"]").click()
                page.locator("input[name=\"BMH\"]").fill(exam_id)
                page.wait_for_timeout(500)

                page.locator("#MM").click()
                page.locator("#MM").fill(password)
                page.wait_for_timeout(500)

                verify_element = page.query_selector("id=verify")
                verify_element.screenshot(path=self.config["screenshot_dir"] / "verify.png")
                with open(self.config["screenshot_dir"] / "verify.png", "rb") as f:
                    verify_code = self.ocr.classification(f.read())

                page.locator("input[name=\"verifyCode\"]").click()
                page.locator("input[name=\"verifyCode\"]").fill(verify_code)
                page.wait_for_timeout(500)

                page.get_by_role("link", name="查询").click()
                page.wait_for_timeout(1000)

                # 结果处理
                student_dir = self.config["scores_dir"] / f"{exam_id}_{name}"
                student_dir.mkdir(exist_ok=True)
                page.screenshot(path=student_dir / "full_page.png")

                if self.config["overwrite"] or not ws.cell(
                        row=row,
                        column=self.config["column_mapping"]["first_score"]
                ).value:
                    html_content = page.content()
                    temp_html = self.config["html_temp_dir"] / f"temp_{exam_id}.html"

                    with open(temp_html, "w", encoding="utf-8") as f:
                        f.write(html_content)

                    with open(temp_html, "r", encoding="utf-8") as f:
                        score_lines = f.readlines()[62:78]

                    score_file = self.config["selected_html_dir"] / f"{exam_id}.txt"
                    with open(score_file, "w", encoding="utf-8") as f:
                        f.writelines(score_lines)

                    scores = self._extract_scores(score_file)
                    os.remove(temp_html)

                    for idx, score in enumerate(scores):
                        ws.cell(
                            row=row,
                            column=self.config["column_mapping"]["first_score"] + idx,
                            value=score
                        )
                    self.logger.info("%s 成绩已保存", name)
                else:
                    self.logger.info("跳过 %s：成绩已存在", name)

            except Exception as e:
                self.logger.error("处理 %s 时出错：%s", name, str(e), exc_info=True)
            finally:
                context.close()
                browser.close()

    def run(self):

        self.logger.info("系统启动")
        try:
            self.workbook = load_workbook(self.config["excel_path"])
            current_row = self.config["start_row"]

            while current_row <= self.config["max_row"]:
                try:
                    self._process_student(current_row)
                    # 保存逻辑修改
                    if self.config["export_new"]:
                        save_path = Path(self.config["export_path"])
                        self.workbook.save(save_path)
                    else:
                        self.workbook.save(self.config["excel_path"])
                except Exception as e:
                    self.logger.error("第 %d 行处理失败：%s", current_row, str(e))
                finally:
                    current_row += 1

            self.logger.info("全部处理完成")
        except Exception as e:
            self.logger.critical("系统错误：%s", str(e), exc_info=True)
        finally:
            if self.workbook:
                self.workbook.close()


if __name__ == "__main__":
    root = tk.Tk()
    app = ConfigGUI(root)
    root.mainloop()
