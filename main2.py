
from playwright.sync_api import Playwright, sync_playwright     # 导入playwright，自动控制浏览器进行查分操作
# 导入openpyxl，进行表格中已登记内容的读取
from openpyxl import load_workbook
# 导入ddddocr，进行验证码自动化识别
import ddddocr
# 自动创建文件夹
import os
import re

# 进行初始变量的赋值
bmh = None
mm = None
verifycode = None
name = None

# 初始化ocr
ocr = ddddocr.DdddOcr()

sheet_row = 2  # 跳过表头，防止影响数据读取

# 打开存储信息的工作簿（改为存放信息的工作簿）
wb = load_workbook("./excel/报名号.xlsx")
ws = wb.active

# 设置查分url（注意务必将此链接更换为对应的查分链接！！！否则程序将不起作用）
PageURL = "https://cx.shmeea.edu.cn/shmeea/q/shzk2024que6dfurya#"


# 定义提取html中分数的函数
def extract_numbers_from_file(file_path):
    # 创建列表存储分数
    numbers = []

    # 打开txt文件
    with open(file_path, 'r', encoding='utf-8') as file:
        for line in file:
            # 逐行定位分数
            matches = re.findall(r'>(\d+\.?\d*)<', line)
            numbers.extend(matches)
    # 列表返回值
    return numbers


# 查分函数定义
def run(pw: Playwright) -> None:
    global bmh, mm, verifycode, name, sheet_row, ocr  # 全局化变量

    # 将工作簿中信息导入变量
    bmh = str(ws.cell(row=sheet_row, column=3).value)
    mm = str(ws.cell(row=sheet_row, column=5).value)
    name = str(ws.cell(row=sheet_row, column=2).value)

    # 启动浏览器（若将headless设置为True，则无用户界面，电脑还可以进行其他工作）
    browser = pw.chromium.launch(headless=True)
    context = browser.new_context()
    page = context.new_page()

    # 进入查分链接
    page.goto(PageURL)

    page.wait_for_timeout(500)  # 延时

    # 填写准考证号
    page.locator("input[name=\"BMH\"]").click()
    page.locator("input[name=\"BMH\"]").fill(bmh)

    page.wait_for_timeout(500)  # 延时

    # 填写密码
    page.locator("#MM").click()
    page.locator("#MM").fill(mm)

    page.wait_for_timeout(500)  # 延时

    # 将验证码元素截图，保存至所选路径
    screenshot = page.query_selector("id=verify")
    screenshot.screenshot(path="./verify/verify.png")

    # 识别验证码
    image = open("./verify/verify.png", "rb").read()
    result = ocr.classification(image)
    verifycode = result

    # 输入识别验证码结果
    page.locator("input[name=\"verifyCode\"]").click()
    page.locator("input[name=\"verifyCode\"]").fill(verifycode)

    page.wait_for_timeout(500)  # 延时

    # 点击查询按钮
    page.get_by_role("link", name="查询").click()

    page.wait_for_timeout(500)  # 延时

    # 设置成绩保存的路径
    new_folder = bmh + name

    # 获取查询后网页内容
    html = page.content()

    # 将网页内容保存为html
    with open("./html/" + new_folder + ".html", "w", encoding="utf-8") as file:
        file.write(html)
        file.close()

    # 将网页内容截取含有分数的部分
    with open("./html/" + new_folder + ".html", "r", encoding="utf-8") as file:
        lines = file.readlines()
        selected_lines = lines[62:78]

        # 将截取后的部分保存为txt
        with open("./selected_html/" + new_folder + ".txt", "w",
                  encoding="utf-8") as new_file:
            new_file.writelines(selected_lines)
            new_file.close()

        file.close()

    os.remove("./html/" + new_folder + ".html")  # 删除html文件

    # 提取出txt中的分数
    file_path = "./selected_html/" + new_folder + ".txt"
    numbers_list = extract_numbers_from_file(file_path)

    print(numbers_list)

    # 保存分数图片
    page.screenshot(path='./scores' + '/' + new_folder + '/' + new_folder + '.png')

    page.wait_for_timeout(500)  # 延时

    if ws.cell(row=sheet_row, column=6).value is None:
        # 保存分数到excel中
        row_num = sheet_row
        start_col = 6

        for i, value in enumerate(numbers_list):
            ws.cell(row=row_num, column=start_col + i, value=value)

        print("已经保存" + name + "成绩！")

    else:
        print("已有成绩,进入下一位")

    # 关闭浏览器
    context.close()
    browser.close()


# 主程序
with sync_playwright() as playwright:
    while sheet_row <= 271:  # 限制在最大行数内（需将数字改为表格的最大行数）

        if ws.cell(row=sheet_row, column=5).value is not None:  # 识别表格中是否有考生的密码

            run(playwright)  # 如有，则运行查分函数
            wb.save("./excel/报名号.xlsx")

            sheet_row += 1  # 进入下一行

        else:

            print("未获得该考生密码，自动进入下一位")
            sheet_row += 1  # 如没有，自动跳转到下一行考生的信息

print('运行结束')  # 运行结束（废话（？
