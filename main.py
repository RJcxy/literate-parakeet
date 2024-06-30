
# 导入playwright，自动控制浏览器进行查分操作
from playwright.sync_api import Playwright, sync_playwright, expect
# 导入openpyxl，进行表格中已登记内容的读取
from openpyxl import Workbook, load_workbook
# 导入ddddocr，进行验证码自动化识别
import ddddocr
# 自动创建文件夹
import os

# 进行初始变量的赋值
zkzh = None
mm = None
verifycode = None
name = None

ocr = ddddocr.DdddOcr()

sheetrow = 2            # 跳过表头，防止影响数据读取

# 打开存储信息的工作簿（改为存放信息的工作簿）
wb = load_workbook("./baoming/报名号.xlsx")
ws = wb.active

# 设置查分url（注意务必将此链接更换为对应的查分链接！！！否则程序将不起作用）
PageURL = "https://cx.shmeea.edu.cn/shmeea/q/sh2024gkcjquery3zd9f"


# 查分函数定义
def run(playwright: Playwright) -> None:

    global zkzh, mm, verifycode, name, sheetrow, ocr         # 全局化变量

    # 将工作簿中信息导入变量
    zkzh = str(ws.cell(row=sheetrow, column=3).value)
    mm = str(ws.cell(row=sheetrow, column=5).value)
    name = str(ws.cell(row=sheetrow, column=2).value)

    # 启动浏览器（若将headless设置为True，则无用户界面，电脑还可以进行其他工作）
    browser = playwright.chromium.launch(headless=False)
    context = browser.new_context()
    page = context.new_page()

    # 进入查分链接
    page.goto(PageURL)

    page.wait_for_timeout(500)         # 延时

    # 填写准考证号
    page.locator("input[name=\"ZKZH\"]").click()
    page.locator("input[name=\"ZKZH\"]").fill(zkzh)

    page.wait_for_timeout(500)         # 延时

    # 填写密码
    page.locator("#MM").click()
    page.locator("#MM").fill(mm)

    page.wait_for_timeout(500)         # 延时

    # 将验证码元素截图，保存至所选路径
    screenshot = page.query_selector("id=verify")
    screenshot.screenshot(path="./verify/verify.png")

    # 识别验证码
    image = open("./verify/verify.png", "rb").read()
    result = ocr.classification(image)
    verifycode = result

    # 输入识别验证码结果
    page.locator("input[name=\"verifyCode\"]").click()
    page.locator("input[name=\"verifyCode\"]").fill(verifycode)

    page.wait_for_timeout(500)         # 延时

    # 点击查询按钮
    page.get_by_role("link", name="查询").click()

    page.wait_for_timeout(1000)         # 延时

    # 设置成绩保存的路径
    new_folder = zkzh + name

    # 保存分数图片
    page.screenshot(path='./scores' + '/' + new_folder + '/' + new_folder + '.png')

    page.wait_for_timeout(500)         # 延时

    # 关闭浏览器
    context.close()
    browser.close()


# 主程序
with sync_playwright() as playwright:

    while sheetrow <= 4:          # 限制在最大行数内（需将数字改为表格的最大行数）

        if ws.cell(row=sheetrow, column=5).value != None:       # 识别表格中是否有考生的密码

            run(playwright)         # 如有，则运行查分函数

            print("已经保存"+name+"成绩！")
            sheetrow += 1           # 进入下一行

        else:

            print("未获得该考生密码，自动进入下一位")
            sheetrow += 1           # 如没有，自动跳转到下一行考生的信息

print('运行结束')                    # 运行结束（废话（？
